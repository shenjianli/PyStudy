import openpyxl





def get_average_value(file_name):

    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook['Sheet1']

    time_column_index = -1
    value_column_index = -1
    for index in range(sheet.max_column):
        data_value = sheet.cell(row=1, column=index + 1).value
        if '时间' in str(data_value):
            time_column_index = index + 1
        if '数值' in str(data_value):
            value_column_index = index + 1
    sum = 0
    num = 0
    for index in range(1, sheet.max_row):
        data_value = sheet.cell(row=index + 1, column=time_column_index).value
        hour_str = str(data_value)
        hour_str = hour_str[hour_str.rfind('.') + 1:]
        hour_int = int(hour_str)
        if 10 <= hour_int <= 18:
            int_data_value = sheet.cell(row=index + 1, column=value_column_index).value
            sum += int(int_data_value)
            num += 1
    if num == 0:
        num = 1
    ager_value = sum / num
    print(file_name,'平局值：', ager_value)
    return ager_value


def get_titles(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook['Sheet1']
    titles = []
    for index in range(sheet.max_column):
        data_value = sheet.cell(row=1, column=index + 1).value
        if '数值' in str(data_value):
            titles.append('平均值')
        else:
            titles.append(data_value)
    return titles

def get_start_time_unit(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook['Sheet1']

    time = ""
    unit = ""
    time_column_index = -1
    value_column_index = -1
    for index in range(sheet.max_column):
        data_value = sheet.cell(row=1, column=index + 1).value
        if '时间' in str(data_value):
            time_column_index = index + 1
        if '单位' in str(data_value):
            value_column_index = index + 1

    for index in range(1, sheet.max_row):
        data_value = sheet.cell(row=index + 1, column=time_column_index).value
        hour_str = str(data_value)
        hour_str = hour_str[hour_str.rfind('.') + 1:]
        hour_int = int(hour_str)
        if 10 == hour_int:
            int_data_value = sheet.cell(row=index + 1, column=value_column_index).value
            time = str(data_value)
            unit = str(int_data_value)
    return time,unit

def start(files):
    if len(files) > 0:
        wb = openpyxl.Workbook()
        sheet = wb.active
        temp_file = files[0]
        titles = get_titles(temp_file)
        print(titles)
        sheet.append(titles)

        time,unit = get_start_time_unit(temp_file)
        print(time,unit)


        for index in range(len(files)):
            data = []
            file_name = files[index]
            id = index + 1
            wei_id = file_name[file_name.rfind('/')+1:file_name.rfind('.')]
            average_value = get_average_value(file_name)
            data.append(id)
            data.append(wei_id)
            data.append(unit)
            data.append(average_value)
            data.append(time)
            sheet.append(data)

        wb.save('./file/result.xlsx')


if __name__ == '__main__':
    files = ['./file/FI10001.xlsx', './file/FI10002.xlsx','./file/FI10003.xlsx']
    start(files)